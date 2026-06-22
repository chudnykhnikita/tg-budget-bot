import os
import re
import json
import logging
import asyncio
import tempfile
import uuid
import warnings
from datetime import datetime, date, timedelta
from io import BytesIO
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from zoneinfo import ZoneInfo
from dotenv import load_dotenv

load_dotenv()

from telegram import Update, ReplyKeyboardMarkup, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.error import NetworkError, TimedOut
from telegram.warnings import PTBUserWarning

warnings.filterwarnings("ignore", category=PTBUserWarning)

from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ConversationHandler,
    CallbackQueryHandler,
    filters,
    ContextTypes,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Состояния диалога
# ---------------------------------------------------------------------------
(
    ST_CHOOSE_ACCOUNT,            # выбор счёта: Наличные / Карта
    ST_CHOOSE_DIRECTION,          # выбор: Поступление / Списание
    ST_ENTERING_AMOUNT,           # ввод суммы
    ST_CHOOSE_CATEGORY,           # выбор категории кнопками (или свободный ввод)
    ST_ENTERING_NOTE,             # ввод примечания 2-го уровня
    ST_ENTERING_ZP_DATE,          # ввод даты для ЗП упаковщиков
    ST_ENTERING_REQUEST_AMOUNT,   # ввод суммы запроса
    ST_CHOOSE_TRANSFER_DIR,       # выбор направления перевода (карта↔наличные)
    ST_ENTERING_TRANSFER_AMOUNT,  # ввод суммы перевода
    ST_CHOOSE_BANK,               # выбор банка для карты / запроса
    ST_CHOOSE_TRANSFER_BANK_TO,   # выбор банка-получателя при переводе карта→карта
) = range(11)

(
    EDIT_CHOOSE_TYPE,       # выбор: Наличные / Карта / Запросы
    EDIT_CHOOSE_DIRECTION,  # выбор: Поступления / Списания
    EDIT_LIST,              # пагинированный список операций или запросов
    EDIT_CHOOSE_FIELD,      # выбор поля для редактирования или удаления
    EDIT_ENTERING_VALUE,    # ввод нового значения
    EDIT_CONFIRM_DELETE,    # подтверждение удаления
) = range(10, 16)

EDIT_PAGE_SIZE = 5

# ---------------------------------------------------------------------------
# Константы
# ---------------------------------------------------------------------------
DATA_FILE       = os.environ.get("DATA_FILE", "/data/budget_data.json")
MOSCOW_TZ       = ZoneInfo("Europe/Moscow")
MAX_NOTE_LEN    = 128

# Категории (константы — используются и как лейблы кнопок, и как ключи).
CAT_BANK       = "🏦 Плата банку"
CAT_ZP         = "👷 ЗП упаковщиков"
CAT_ZP_WORKERS = "💼 ЗП работников"
CAT_DIV        = "💎 Дивиденды"
CAT_MEAL       = "🍽 Обеды"
CAT_OFC        = "🏢 Офис"
CAT_WH         = "📦 Склад"

CAT_SALE  = "🛒 Продажа со склада"

SUB_BANK_SERVICE = "🧾 Обслуживание счёта"
SUB_BANK_FEE     = "💸 Комиссия"

CAT_COMP = "↩️ Компенсация клиенту за брак"

# Категории карты — списания
CARD_EXPENSE_CATEGORIES = [CAT_BANK, CAT_ZP, CAT_ZP_WORKERS, CAT_DIV, CAT_MEAL, CAT_OFC, CAT_WH, CAT_COMP]
# Подкатегории «Плата банку»
BANK_SUBCATEGORIES = [SUB_BANK_SERVICE, SUB_BANK_FEE]
# Подкатегории «Дивиденды» (имена без эмодзи)
DIVIDEND_SUBCATEGORIES = ["Андрей", "Алексей", "Никита"]
# Подкатегории «ЗП работников» — выплаты по частям
WORKERS_SUBCATEGORIES = ["Никита 1 часть", "Никита 2 часть", "Михаил 1 часть", "Михаил 2 часть", "Премия Михаил", "Премия Никита"]

# Категории наличных — поступления (кнопки + свободный ввод)
CASH_INCOME_CATEGORIES = [CAT_SALE]

# Категории наличных — списания (кнопки + свободный ввод)
CASH_EXPENSE_CATEGORIES = [CAT_ZP]

BTN_FREE              = "✏️ Свой вариант"
BTN_TODAY             = "📅 Сегодня"
BTN_YESTERDAY         = "📅 Вчера"
BTN_BEFORE_YESTERDAY  = "📅 Позавчера"
BTN_CANCEL            = "❌ Отмена"

# Банки
BANK_TINKOFF     = "tinkoff"
BANK_VTB         = "vtb"
BTN_BANK_TINKOFF = "🟡 Тиньков"
BTN_BANK_VTB     = "🔵 ВТБ"
BANK_BUTTONS     = [BTN_BANK_TINKOFF, BTN_BANK_VTB]
BANK_LABELS      = {BANK_TINKOFF: "Тиньков", BANK_VTB: "ВТБ"}


def _bank_from_btn(text: str) -> str | None:
    if text == BTN_BANK_TINKOFF:
        return BANK_TINKOFF
    if text == BTN_BANK_VTB:
        return BANK_VTB
    return None


def _bank_label(bank: str | None) -> str:
    return BANK_LABELS.get(bank or "", "—")

# ---------------------------------------------------------------------------
# Главная клавиатура
# ---------------------------------------------------------------------------
MAIN_KEYBOARD = ReplyKeyboardMarkup(
    [
        ["💵 Наличные", "💳 Карта"],
        ["🔄 Перевод"],
        ["💰 Баланс",   "🕓 История"],
        ["✏️ Изменить", "📨 Запросил"],
    ],
    resize_keyboard=True,
    is_persistent=True,
)

_lock = asyncio.Lock()

# ---------------------------------------------------------------------------
# Работа с данными
# ---------------------------------------------------------------------------

def _empty_data() -> dict:
    return {"transactions": [], "requests": [], "transfers": []}


def _migrate(data: dict) -> dict:
    """Идемпотентная миграция данных к актуальной схеме."""
    data.setdefault("transactions", [])
    data.setdefault("requests", [])
    data.setdefault("transfers", [])

    # Транзакции до cash/card-разделения: дополняем id/account/note/bank,
    # подстраховываемся от отсутствующего type, нормализуем amount к "X.XX".
    for t in data["transactions"]:
        if "id" not in t:
            t["id"] = str(uuid.uuid4())
        if "account" not in t:
            t["account"] = "card"   # legacy default — большинство операций были по карте
        if "note" not in t:
            t["note"] = None
        if "type" not in t:
            t["type"] = "expense"   # подстраховка
        # Все старые карточные операции — Тиньков (ВТБ появился только 2026-05-27)
        if "bank" not in t:
            t["bank"] = BANK_TINKOFF if t.get("account") == "card" else None
        amt = t.get("amount", 0)
        try:
            t["amount"] = str(Decimal(str(amt)).quantize(Decimal("0.01"), ROUND_HALF_UP))
        except (InvalidOperation, ValueError):
            t["amount"] = "0.00"

    # Запросы и переводы — id/amount нормализация на всякий случай
    for r in data["requests"]:
        if "id" not in r:
            r["id"] = str(uuid.uuid4())
        # Все старые запросы — Тиньков
        if "bank" not in r:
            r["bank"] = BANK_TINKOFF
        try:
            r["amount"] = str(Decimal(str(r.get("amount", 0))).quantize(Decimal("0.01"), ROUND_HALF_UP))
        except (InvalidOperation, ValueError):
            r["amount"] = "0.00"
    for t in data["transfers"]:
        if "id" not in t:
            t["id"] = str(uuid.uuid4())
        # Все старые переводы — Тиньков
        if "bank" not in t:
            t["bank"] = BANK_TINKOFF
        try:
            t["amount"] = str(Decimal(str(t.get("amount", 0))).quantize(Decimal("0.01"), ROUND_HALF_UP))
        except (InvalidOperation, ValueError):
            t["amount"] = "0.00"

    return data


def load_data() -> dict:
    if not os.path.exists(DATA_FILE):
        return _empty_data()
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
    except json.JSONDecodeError:
        # Бэкап с таймстампом — чтобы не перетирать прошлые повреждения
        ts = datetime.now(MOSCOW_TZ).strftime("%Y%m%d_%H%M%S")
        backup = f"{DATA_FILE}.corrupt.{ts}"
        logger.exception("Corrupted %s, backing up to %s", DATA_FILE, backup)
        try:
            os.rename(DATA_FILE, backup)
        except OSError:
            logger.exception("Failed to backup corrupt data file")
        return _empty_data()
    return _migrate(data)


def _save_data_sync(data: dict):
    """Синхронная атомарная запись через временный файл + os.replace.
    НЕ берёт лок — для использования из под уже захваченного _lock."""
    dir_ = os.path.dirname(DATA_FILE) or "."
    os.makedirs(dir_, exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=dir_, prefix=".budget_", suffix=".json")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        os.replace(tmp, DATA_FILE)
    except Exception:
        try:
            os.unlink(tmp)
        except FileNotFoundError:
            pass
        raise


async def save_data(data: dict):
    """Атомарная запись под глобальным локом."""
    async with _lock:
        await asyncio.to_thread(_save_data_sync, data)


async def _atomic_modify(mutator):
    """Атомарно: load → mutator(data) → save. Под одним локом —
    защищает от гонок между конкурентными write-операциями."""
    async with _lock:
        data = await asyncio.to_thread(load_data)
        mutator(data)
        await asyncio.to_thread(_save_data_sync, data)


async def _save_transaction(user_id: int, amount: Decimal, t_type: str,
                             account: str, category: str | None, note: str | None,
                             bank: str | None = None):
    now = datetime.now(MOSCOW_TZ).strftime("%Y-%m-%d %H:%M")
    record = {
        "id":       str(uuid.uuid4()),
        "user_id":  user_id,
        "type":     t_type,        # "income" | "expense"
        "account":  account,       # "cash" | "card"
        "amount":   str(amount),
        "category": category,
        "note":     note,
        "bank":     bank,          # "tinkoff" | "vtb" | None (наличные)
        "date":     now,
    }
    await _atomic_modify(lambda data: data["transactions"].append(record))


async def _save_request(user_id: int, amount: Decimal, bank: str | None = None):
    now = datetime.now(MOSCOW_TZ).strftime("%Y-%m-%d %H:%M")
    record = {
        "id":      str(uuid.uuid4()),
        "user_id": user_id,
        "amount":  str(amount),
        "bank":    bank,
        "date":    now,
    }
    await _atomic_modify(lambda data: data["requests"].append(record))


async def _save_transfer(user_id: int, amount: Decimal, src: str, dst: str,
                          bank: str | None = None, bank_to: str | None = None):
    """src/dst: 'cash' или 'card'. bank — банк-отправитель, bank_to — банк-получатель (карта→карта)."""
    now = datetime.now(MOSCOW_TZ).strftime("%Y-%m-%d %H:%M")
    record = {
        "id":      str(uuid.uuid4()),
        "user_id": user_id,
        "amount":  str(amount),
        "from":    src,
        "to":      dst,
        "bank":    bank,
        "bank_to": bank_to,
        "date":    now,
    }
    await _atomic_modify(lambda data: data["transfers"].append(record))


def parse_amount(s: str) -> Decimal:
    try:
        v = Decimal(s.strip().replace(",", ".")).quantize(Decimal("0.01"), ROUND_HALF_UP)
    except InvalidOperation:
        raise ValueError
    if v <= 0:
        raise ValueError
    return v


def fmt(amount) -> str:
    """Форматирует сумму для отображения в боте: целое число с пробельными
    разделителями тысяч. Копейки округляются (ROUND_HALF_UP).
    Пример: Decimal('10000.00') → '10 000'.
    Для Excel НЕ используется — там значения идут как float напрямую."""
    d = Decimal(str(amount)).quantize(Decimal("1"), ROUND_HALF_UP)
    return f"{d:,}".replace(",", " ")  # неразрывный пробел — чтобы число не разрывалось переносом

# Регулярка для эмодзи: основные блоки Misc/Pictographs, Emoticons, Transport,
# Supplemental, Misc Symbols + Dingbats и selector-16. Используется только при
# экспорте в Excel — в JSON и в боте эмодзи остаются на своих местах.
_EMOJI_RE = re.compile(
    "["
    "\U0001F300-\U0001F5FF"   # символы и пиктограммы
    "\U0001F600-\U0001F64F"   # эмоции
    "\U0001F680-\U0001F6FF"   # транспорт / карта (включая 🛒)
    "\U0001F900-\U0001F9FF"   # supplemental (включая 🧾)
    "☀-➿"           # misc symbols + dingbats (✏️, ❌, ◀️ и т.п.)
    "️"                   # variation selector-16
    "]+"
)


def _strip_emoji(s) -> str:
    if not s:
        return ""
    return _EMOJI_RE.sub("", s).strip()


# ---------------------------------------------------------------------------
# Вспомогательные функции — клавиатуры
# ---------------------------------------------------------------------------

def _reply_kb(options: list[str], add_free: bool = False, add_dates: bool = False,
              cols: int = 2) -> ReplyKeyboardMarkup:
    # Раскладываем варианты сеткой по `cols` колонок (по умолчанию 2),
    # как основная клавиатура.
    rows = [options[i:i + cols] for i in range(0, len(options), cols)]
    if add_dates:
        rows.append([BTN_TODAY, BTN_YESTERDAY, BTN_BEFORE_YESTERDAY])
    if add_free:
        rows.append([BTN_FREE])
    rows.append([BTN_CANCEL])
    return ReplyKeyboardMarkup(rows, resize_keyboard=True, one_time_keyboard=True)

# ---------------------------------------------------------------------------
# Error handler
# ---------------------------------------------------------------------------

async def on_error(update: object, context: ContextTypes.DEFAULT_TYPE):
    err = context.error
    if isinstance(err, (NetworkError, TimedOut)):
        logger.warning("Transient network error: %s", err)
        return
    logger.exception("Unhandled error", exc_info=err)
    if isinstance(update, Update) and update.effective_message:
        await update.effective_message.reply_text("⚠️ Что-то пошло не так, попробуй ещё раз.")

# ---------------------------------------------------------------------------
# /start
# ---------------------------------------------------------------------------

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 Привет! Я бот для ведения бюджета.\n\n"
        "Выбери счёт кнопками внизу.\n\n"
        "/clear — очистить все данные",
        reply_markup=MAIN_KEYBOARD,
    )
    return ConversationHandler.END

# ---------------------------------------------------------------------------
# Шаг 1 — выбор счёта (Наличные / Карта)
# ---------------------------------------------------------------------------

async def handle_account(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text or ""

    if "Баланс" in text:
        await show_summary(update, context)
        return ConversationHandler.END
    if "История" in text:
        await history(update, context)
        return ConversationHandler.END

    if "Запросил" in text:
        context.user_data["flow"] = "request"
        await update.message.reply_text(
            "Введи запрошенную сумму:",
            reply_markup=ReplyKeyboardMarkup(
                [[BTN_CANCEL]], resize_keyboard=True, one_time_keyboard=True
            ),
        )
        return ST_ENTERING_REQUEST_AMOUNT

    if "Перевод" in text:
        kb = ReplyKeyboardMarkup(
            [
                ["💳 → 💵  Карта → Наличные"],
                ["💵 → 💳  Наличные → Карта"],
                ["💳 → 💳  Карта → Карта"],
                [BTN_CANCEL],
            ],
            resize_keyboard=True, one_time_keyboard=True,
        )
        await update.message.reply_text("Куда переводим?", reply_markup=kb)
        return ST_CHOOSE_TRANSFER_DIR

    if "Наличные" in text:
        context.user_data["account"] = "cash"
        kb = ReplyKeyboardMarkup(
            [["➕ Поступление", "➖ Списание"], [BTN_CANCEL]],
            resize_keyboard=True, one_time_keyboard=True,
        )
        await update.message.reply_text("Поступление или списание?", reply_markup=kb)
        return ST_CHOOSE_DIRECTION
    elif "Карта" in text:
        context.user_data["account"] = "card"
        context.user_data["bank"] = BANK_TINKOFF
        kb = ReplyKeyboardMarkup(
            [["➕ Поступление", "➖ Списание"], [BTN_CANCEL]],
            resize_keyboard=True, one_time_keyboard=True,
        )
        await update.message.reply_text("Поступление или списание?", reply_markup=kb)
        return ST_CHOOSE_DIRECTION
    else:
        return ConversationHandler.END


# ---------------------------------------------------------------------------
# Шаг 1б — выбор банка (Тиньков / ВТБ)
# ---------------------------------------------------------------------------

async def handle_bank(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (update.message.text or "").strip()
    if BTN_CANCEL in text:
        return await cancel(update, context)

    bank = _bank_from_btn(text)
    if bank is None:
        kb = _reply_kb(BANK_BUTTONS)
        await update.message.reply_text("Выбери банк из предложенных:", reply_markup=kb)
        return ST_CHOOSE_BANK

    context.user_data["bank"] = bank
    flow = context.user_data.get("flow")

    if flow == "transfer":
        tr = context.user_data.get("transfer", {})
        if tr.get("from") == "card" and tr.get("to") == "card":
            # Карта→Карта: выбираем банк-получатель
            kb = _reply_kb(BANK_BUTTONS)
            await update.message.reply_text("Банк-получатель?", reply_markup=kb)
            return ST_CHOOSE_TRANSFER_BANK_TO
        await update.message.reply_text(
            "Введи сумму перевода:",
            reply_markup=ReplyKeyboardMarkup(
                [[BTN_CANCEL]], resize_keyboard=True, one_time_keyboard=True
            ),
        )
        return ST_ENTERING_TRANSFER_AMOUNT

    # По умолчанию — поток «Карта», переходим к выбору направления
    kb = ReplyKeyboardMarkup(
        [["➕ Поступление", "➖ Списание"], [BTN_CANCEL]],
        resize_keyboard=True, one_time_keyboard=True,
    )
    await update.message.reply_text("Поступление или списание?", reply_markup=kb)
    return ST_CHOOSE_DIRECTION

# ---------------------------------------------------------------------------
# Шаг 2 — выбор направления
# ---------------------------------------------------------------------------

async def handle_direction(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text or ""
    if BTN_CANCEL in text:
        return await cancel(update, context)

    if "Поступление" in text:
        context.user_data["direction"] = "income"
    elif "Списание" in text:
        context.user_data["direction"] = "expense"
    else:
        return ST_CHOOSE_DIRECTION

    await update.message.reply_text("Введи сумму:", reply_markup=ReplyKeyboardMarkup(
        [[BTN_CANCEL]], resize_keyboard=True, one_time_keyboard=True
    ))
    return ST_ENTERING_AMOUNT

# ---------------------------------------------------------------------------
# Шаг 3 — ввод суммы
# ---------------------------------------------------------------------------

async def handle_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (update.message.text or "").strip()
    if BTN_CANCEL in text:
        return await cancel(update, context)

    try:
        amount = parse_amount(text)
    except ValueError:
        await update.message.reply_text("❌ Введи корректную сумму (например: 1500 или 99.90)")
        return ST_ENTERING_AMOUNT

    context.user_data["amount"] = str(amount)
    return await _go_to_category(update, context)

# ---------------------------------------------------------------------------
# Шаг 4 — выбор категории
# ---------------------------------------------------------------------------

async def _go_to_category(update: Update, context: ContextTypes.DEFAULT_TYPE):
    account   = context.user_data["account"]
    direction = context.user_data["direction"]

    # Наличные — поступление
    if account == "cash" and direction == "income":
        kb = _reply_kb(CASH_INCOME_CATEGORIES, add_free=True)
        await update.message.reply_text("Выбери категорию:", reply_markup=kb)
        return ST_CHOOSE_CATEGORY

    # Наличные — списание
    if account == "cash" and direction == "expense":
        kb = _reply_kb(CASH_EXPENSE_CATEGORIES, add_free=True)
        await update.message.reply_text("Выбери категорию:", reply_markup=kb)
        return ST_CHOOSE_CATEGORY

    # Карта — поступление (без категории)
    if account == "card" and direction == "income":
        return await _finish(update, context, category=None, note=None)

    # Карта — списание
    if account == "card" and direction == "expense":
        kb = _reply_kb(CARD_EXPENSE_CATEGORIES, add_free=True)
        await update.message.reply_text("Выбери категорию:", reply_markup=kb)
        return ST_CHOOSE_CATEGORY

    return ConversationHandler.END


async def handle_category(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (update.message.text or "").strip()
    if BTN_CANCEL in text:
        return await cancel(update, context)

    account   = context.user_data["account"]
    direction = context.user_data["direction"]

    # Свободный ввод
    if text == BTN_FREE:
        await update.message.reply_text(
            "Напиши свою категорию:",
            reply_markup=ReplyKeyboardMarkup([[BTN_CANCEL]], resize_keyboard=True, one_time_keyboard=True)
        )
        context.user_data["awaiting_free_category"] = True
        return ST_CHOOSE_CATEGORY

    if context.user_data.pop("awaiting_free_category", False):
        category = text[:MAX_NOTE_LEN]
        context.user_data["category"] = category
        return await _finish(update, context, category=category, note=None)

    # Карта — списание: подкатегории
    if account == "card" and direction == "expense":
        if context.user_data.pop("awaiting_subcategory", False):
            context.user_data["note"] = text[:MAX_NOTE_LEN]
            return await _finish(update, context,
                                  category=context.user_data["category"],
                                  note=context.user_data["note"])

        if text == CAT_BANK:
            context.user_data["category"] = text
            kb = _reply_kb(BANK_SUBCATEGORIES)
            await update.message.reply_text("Уточни:", reply_markup=kb)
            context.user_data["awaiting_subcategory"] = True
            return ST_CHOOSE_CATEGORY

        if text == CAT_DIV:
            context.user_data["category"] = text
            kb = _reply_kb(DIVIDEND_SUBCATEGORIES)
            await update.message.reply_text("Кому?", reply_markup=kb)
            context.user_data["awaiting_subcategory"] = True
            return ST_CHOOSE_CATEGORY

        if text == CAT_ZP_WORKERS:
            context.user_data["category"] = text
            kb = _reply_kb(WORKERS_SUBCATEGORIES)
            await update.message.reply_text("Кому и какая часть?", reply_markup=kb)
            context.user_data["awaiting_subcategory"] = True
            return ST_CHOOSE_CATEGORY

        if text == CAT_ZP:
            context.user_data["category"] = text
            kb = _reply_kb([], add_dates=True)
            await update.message.reply_text("Укажи дату выплаты (или нажми «Сегодня»):", reply_markup=kb)
            return ST_ENTERING_ZP_DATE

        if text in (CAT_OFC, CAT_WH):
            context.user_data["category"] = text
            await update.message.reply_text(
                "Добавь пояснение (или /skip чтобы пропустить):",
                reply_markup=ReplyKeyboardMarkup([[BTN_CANCEL]], resize_keyboard=True, one_time_keyboard=True)
            )
            return ST_ENTERING_NOTE

        if text == CAT_MEAL:
            return await _finish(update, context, category=text, note=None)

        if text in CARD_EXPENSE_CATEGORIES:
            context.user_data["category"] = text
            return await _finish(update, context, category=text, note=None)

    # Наличные — ЗП упаковщиков
    if text == CAT_ZP:
        context.user_data["category"] = text
        kb = _reply_kb([], add_dates=True)
        await update.message.reply_text("Укажи дату выплаты (или нажми «Сегодня»):", reply_markup=kb)
        return ST_ENTERING_ZP_DATE

    # Все остальные кнопки — сохраняем как категорию без доп. ввода
    context.user_data["category"] = text[:MAX_NOTE_LEN]
    return await _finish(update, context, category=context.user_data["category"], note=None)


async def handle_zp_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (update.message.text or "").strip()
    if BTN_CANCEL in text:
        return await cancel(update, context)

    today = date.today()
    if text == BTN_TODAY:
        note = today.strftime("%d.%m.%Y")
    elif text == BTN_YESTERDAY:
        note = (today - timedelta(days=1)).strftime("%d.%m.%Y")
    elif text == BTN_BEFORE_YESTERDAY:
        note = (today - timedelta(days=2)).strftime("%d.%m.%Y")
    else:
        note = text[:MAX_NOTE_LEN]

    return await _finish(update, context,
                          category=context.user_data["category"],
                          note=note)


async def handle_note(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (update.message.text or "").strip()
    if BTN_CANCEL in text:
        return await cancel(update, context)
    note = None if text == "/skip" else text[:MAX_NOTE_LEN]
    return await _finish(update, context,
                          category=context.user_data.get("category"),
                          note=note)

# ---------------------------------------------------------------------------
# Финальное сохранение
# ---------------------------------------------------------------------------

async def _finish(update: Update, context: ContextTypes.DEFAULT_TYPE,
                   category: str | None, note: str | None):
    amount    = Decimal(context.user_data["amount"])
    account   = context.user_data["account"]
    direction = context.user_data["direction"]
    bank      = context.user_data.get("bank")  # None для наличных

    try:
        await _save_transaction(
            user_id=update.effective_user.id,
            amount=amount,
            t_type=direction,
            account=account,
            category=category,
            note=note,
            bank=bank,
        )
    except OSError:
        logger.exception("Failed to save transaction")
        await update.message.reply_text(
            "⚠️ Не удалось сохранить операцию, данные НЕ записаны.",
            reply_markup=MAIN_KEYBOARD,
        )
        return ConversationHandler.END

    acc_label  = "💵 Наличные" if account == "cash" else "💳 Карта"
    bank_label = f" [{_bank_label(bank)}]" if bank else ""
    dir_label  = "Поступление" if direction == "income" else "Списание"
    cat_label  = f" · {category}" if category else ""
    note_label = f" · {note}" if note else ""

    await update.message.reply_text(
        f"✅ {acc_label}{bank_label} | {dir_label} {fmt(amount)} ₽{cat_label}{note_label} сохранено!",
        reply_markup=MAIN_KEYBOARD,
    )
    context.user_data.clear()
    await show_summary(update, context)
    return ConversationHandler.END

# ---------------------------------------------------------------------------
# Запрос денег («Запросил»)
# ---------------------------------------------------------------------------

async def handle_request_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (update.message.text or "").strip()
    if BTN_CANCEL in text:
        return await cancel(update, context)

    try:
        amount = parse_amount(text)
    except ValueError:
        await update.message.reply_text("❌ Введи корректную сумму (например: 100000 или 99.90)")
        return ST_ENTERING_REQUEST_AMOUNT

    user_id = update.effective_user.id
    try:
        await _save_request(user_id, amount)
    except OSError:
        logger.exception("Failed to save request")
        await update.message.reply_text(
            "⚠️ Не удалось сохранить запрос, попробуй ещё раз.",
            reply_markup=MAIN_KEYBOARD,
        )
        return ConversationHandler.END

    # Показываем сводку: сколько всего запрошено, сколько уже пришло, сколько осталось
    data    = await asyncio.to_thread(load_data)
    txs     = [t for t in data["transactions"] if t["user_id"] == user_id]
    reqs    = [r for r in data["requests"]     if r["user_id"] == user_id]
    total_req     = sum(Decimal(r["amount"]) for r in reqs)
    total_card_in = sum(
        Decimal(t["amount"]) for t in txs
        if t.get("account") == "card" and t["type"] == "income"
    )
    remaining = total_req - total_card_in

    await update.message.reply_text(
        f"✅ Запрошено {fmt(amount)} ₽\n\n"
        f"📨 Всего запрошено:  {fmt(total_req)} ₽\n"
        f"💳 Получено по карте: {fmt(total_card_in)} ₽\n"
        f"⏳ Осталось получить: {fmt(remaining)} ₽",
        reply_markup=MAIN_KEYBOARD,
    )
    context.user_data.clear()
    await show_summary(update, context)
    return ConversationHandler.END

# ---------------------------------------------------------------------------
# Переводы между счетами
# ---------------------------------------------------------------------------

async def handle_transfer_dir(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text or ""
    if BTN_CANCEL in text:
        return await cancel(update, context)

    if "Карта → Наличные" in text:
        context.user_data["transfer"] = {"from": "card", "to": "cash"}
    elif "Наличные → Карта" in text:
        context.user_data["transfer"] = {"from": "cash", "to": "card"}
    elif "Карта → Карта" in text:
        context.user_data["transfer"] = {"from": "card", "to": "card"}
    else:
        return ST_CHOOSE_TRANSFER_DIR

    context.user_data["flow"] = "transfer"
    context.user_data["bank"] = BANK_TINKOFF
    if context.user_data["transfer"]["from"] == "card" and context.user_data["transfer"]["to"] == "card":
        context.user_data["bank_to"] = BANK_TINKOFF
    await update.message.reply_text(
        "Введи сумму перевода:",
        reply_markup=ReplyKeyboardMarkup(
            [[BTN_CANCEL]], resize_keyboard=True, one_time_keyboard=True
        ),
    )
    return ST_ENTERING_TRANSFER_AMOUNT


async def handle_transfer_bank_to(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Выбор банка-получателя при переводе карта→карта."""
    text = (update.message.text or "").strip()
    if BTN_CANCEL in text:
        return await cancel(update, context)

    bank_to = _bank_from_btn(text)
    if bank_to is None:
        kb = _reply_kb(BANK_BUTTONS)
        await update.message.reply_text("Выбери банк из предложенных:", reply_markup=kb)
        return ST_CHOOSE_TRANSFER_BANK_TO

    bank_from = context.user_data.get("bank")
    if bank_to == bank_from:
        kb = _reply_kb(BANK_BUTTONS)
        await update.message.reply_text("❌ Банк-получатель должен отличаться от отправителя:", reply_markup=kb)
        return ST_CHOOSE_TRANSFER_BANK_TO

    context.user_data["bank_to"] = bank_to
    await update.message.reply_text(
        "Введи сумму перевода:",
        reply_markup=ReplyKeyboardMarkup(
            [[BTN_CANCEL]], resize_keyboard=True, one_time_keyboard=True
        ),
    )
    return ST_ENTERING_TRANSFER_AMOUNT


async def handle_transfer_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (update.message.text or "").strip()
    if BTN_CANCEL in text:
        return await cancel(update, context)

    try:
        amount = parse_amount(text)
    except ValueError:
        await update.message.reply_text("❌ Введи корректную сумму (например: 5000 или 99.90)")
        return ST_ENTERING_TRANSFER_AMOUNT

    tr      = context.user_data.get("transfer", {})
    src     = tr.get("from")
    dst     = tr.get("to")
    bank    = context.user_data.get("bank")
    bank_to = context.user_data.get("bank_to")
    if src not in ("cash", "card") or dst not in ("cash", "card"):
        await update.message.reply_text(
            "⚠️ Что-то пошло не так с направлением. Попробуй ещё раз.",
            reply_markup=MAIN_KEYBOARD,
        )
        return ConversationHandler.END

    user_id = update.effective_user.id
    try:
        await _save_transfer(user_id, amount, src, dst, bank=bank, bank_to=bank_to)
    except OSError:
        logger.exception("Failed to save transfer")
        await update.message.reply_text(
            "⚠️ Не удалось сохранить перевод, попробуй ещё раз.",
            reply_markup=MAIN_KEYBOARD,
        )
        return ConversationHandler.END

    src_lbl      = "💳 Карта" if src == "card" else "💵 Наличные"
    dst_lbl      = "💳 Карта" if dst == "card" else "💵 Наличные"
    bank_from_lbl = f" [{_bank_label(bank)}]"    if bank    else ""
    bank_to_lbl   = f" [{_bank_label(bank_to)}]" if bank_to else ""
    await update.message.reply_text(
        f"✅ Перевод {fmt(amount)} ₽\n{src_lbl}{bank_from_lbl} → {dst_lbl}{bank_to_lbl} сохранён.",
        reply_markup=MAIN_KEYBOARD,
    )
    context.user_data.clear()
    await show_summary(update, context)
    return ConversationHandler.END

# ---------------------------------------------------------------------------
# Баланс
# ---------------------------------------------------------------------------

async def show_summary(update: Update, context: ContextTypes.DEFAULT_TYPE):
    data    = await asyncio.to_thread(load_data)
    user_id = update.effective_user.id
    txs     = [t for t in data["transactions"] if t["user_id"] == user_id]
    reqs    = [r for r in data["requests"]     if r["user_id"] == user_id]
    trs     = [t for t in data["transfers"]    if t["user_id"] == user_id]

    if not txs and not reqs and not trs:
        await update.message.reply_text("📭 Операций пока нет.", reply_markup=MAIN_KEYBOARD)
        return

    def totals(account, bank=None):
        """Поступления и списания по счёту; bank=None — все банки."""
        inc = sum(Decimal(t["amount"]) for t in txs
                  if t.get("account") == account and t.get("type") == "income"
                  and (bank is None or t.get("bank") == bank))
        exp = sum(Decimal(t["amount"]) for t in txs
                  if t.get("account") == account and t.get("type") == "expense"
                  and (bank is None or t.get("bank") == bank))
        return inc, exp

    cash_inc, cash_exp = totals("cash")
    card_inc, card_exp = totals("card")

    # Переводы: чистое движение по каждому счёту (переводы не привязаны к банку)
    cash_tr_in  = sum(Decimal(t["amount"]) for t in trs if t["to"]   == "cash")
    cash_tr_out = sum(Decimal(t["amount"]) for t in trs if t["from"] == "cash")
    card_tr_in  = sum(Decimal(t["amount"]) for t in trs if t["to"]   == "card")
    card_tr_out = sum(Decimal(t["amount"]) for t in trs if t["from"] == "card")
    cash_tr_net = cash_tr_in - cash_tr_out
    card_tr_net = card_tr_in - card_tr_out

    cash_bal = cash_inc - cash_exp + cash_tr_net
    card_bal = card_inc - card_exp + card_tr_net
    total    = cash_bal + card_bal

    # Запросы: сколько запрошено и сколько ещё не пришло (минус поступления на карту)
    total_req = sum(Decimal(r["amount"]) for r in reqs)
    remaining = total_req - card_inc
    req_lines = ""
    if reqs:
        req_lines = (
            "\n\n📨 Запросы\n"
            f"  Запрошено:        {fmt(total_req)} ₽\n"
            f"  Получено (карта): {fmt(card_inc)} ₽\n"
            f"  Осталось:         {fmt(remaining)} ₽"
        )

    # Расходы по категориям (карта)
    categories: dict[str, Decimal] = {}
    for t in txs:
        if t["type"] == "expense" and t.get("category"):
            key = t["category"]
            if t.get("note"):
                key += f" · {t['note']}"
            categories[key] = categories.get(key, Decimal("0")) + Decimal(t["amount"])

    cat_lines = ""
    if categories:
        sorted_cats = sorted(categories.items(), key=lambda x: -x[1])
        cat_lines = "\n\n📊 Расходы по категориям:\n" + "\n".join(
            f"  • {k}: {fmt(v)} ₽" for k, v in sorted_cats
        )

    def _signed(v: Decimal) -> str:
        return ("+" + fmt(v)) if v > 0 else fmt(v)

    cash_tr_line = (f"  Переводы:    {_signed(cash_tr_net)} ₽\n" if trs else "")
    card_tr_line = (f"  Переводы:    {_signed(card_tr_net)} ₽\n" if trs else "")

    # Разбивка карты по банкам (показываем только банки у которых есть операции/переводы)
    bank_order = [(BANK_TINKOFF, "🟡 Тиньков"), (BANK_VTB, "🔵 ВТБ")]
    card_bank_lines = ""
    active_banks = 0
    for bank_key, bank_name in bank_order:
        b_inc, b_exp = totals("card", bank=bank_key)
        b_tr_in  = sum(Decimal(t["amount"]) for t in trs
                       if t.get("to") == "card" and (
                           t.get("bank_to") == bank_key if t.get("from") == "card"
                           else t.get("bank") == bank_key
                       ))
        b_tr_out = sum(Decimal(t["amount"]) for t in trs
                       if t.get("from") == "card" and t.get("bank") == bank_key)
        b_tr_net = b_tr_in - b_tr_out
        if b_inc == 0 and b_exp == 0 and b_tr_net == 0:
            continue
        active_banks += 1
        b_bal     = b_inc - b_exp + b_tr_net
        b_tr_line = (f"\n    Переводы:    {_signed(b_tr_net)} ₽" if b_tr_net != 0 else "")
        card_bank_lines += (
            f"\n  {bank_name}\n"
            f"    Поступления: {fmt(b_inc)} ₽\n"
            f"    Списания:    {fmt(b_exp)} ₽"
            f"{b_tr_line}\n"
            f"    Баланс:      {fmt(b_bal)} ₽"
        )

    # Если активных банков больше одного — показываем разбивку и итого
    if active_banks <= 1:
        card_bank_lines = ""
    card_total_line = ""
    if active_banks > 1:
        card_total_line = (
            f"\n  ──────────────────\n"
            f"  Итого по картам\n"
            f"    Поступления: {fmt(card_inc)} ₽\n"
            f"    Списания:    {fmt(card_exp)} ₽\n"
            f"{('    Переводы:    ' + _signed(card_tr_net) + ' ₽\n') if trs else ''}"
            f"    Баланс:      {fmt(card_bal)} ₽"
        )
    else:
        card_total_line = (
            f"\n  Поступления: {fmt(card_inc)} ₽\n"
            f"  Списания:    {fmt(card_exp)} ₽\n"
            f"{card_tr_line}"
            f"  Баланс:      {fmt(card_bal)} ₽"
        )

    text = (
        f"💵 Наличные\n"
        f"  Поступления: {fmt(cash_inc)} ₽\n"
        f"  Списания:    {fmt(cash_exp)} ₽\n"
        f"{cash_tr_line}"
        f"  Баланс:      {fmt(cash_bal)} ₽\n\n"
        f"💳 Карта"
        f"{card_bank_lines}"
        f"{card_total_line}\n\n"
        f"{'✅' if total >= 0 else '⚠️'} Общий баланс: {fmt(total)} ₽"
        f"{req_lines}"
        f"{cat_lines}"
    )

    await update.message.reply_text(text, reply_markup=MAIN_KEYBOARD)

    # Сразу отправляем Excel
    await export_excel(update, context)

# ---------------------------------------------------------------------------
# История
# ---------------------------------------------------------------------------

async def history(update: Update, context: ContextTypes.DEFAULT_TYPE):
    data    = await asyncio.to_thread(load_data)
    user_id = update.effective_user.id
    txs     = [t for t in data["transactions"] if t["user_id"] == user_id]

    if not txs:
        await update.message.reply_text("📭 Операций пока нет.", reply_markup=MAIN_KEYBOARD)
        return

    cash_lines, card_lines = [], []
    for t in txs[-20:][::-1]:
        acc      = t.get("account", "card")
        cat      = f" ({t['category']})" if t.get("category") else ""
        note     = f" · {t['note']}" if t.get("note") else ""
        bank_str = f" [{_bank_label(t.get('bank'))}]" if acc == "card" else ""
        if t.get("type") == "income":
            line = f"➕ +{fmt(t['amount'])} ₽{cat}{note}{bank_str}  [{t['date']}]"
        else:
            line = f"➖ -{fmt(t['amount'])} ₽{cat}{note}{bank_str}  [{t['date']}]"
        if acc == "cash":
            cash_lines.append(line)
        else:
            card_lines.append(line)

    parts = []
    if cash_lines:
        parts.append("💵 Наличные:\n" + "\n".join(cash_lines))
    if card_lines:
        parts.append("💳 Карта:\n" + "\n".join(card_lines))

    await update.message.reply_text(
        "🕓 Последние операции:\n\n" + "\n\n".join(parts),
        reply_markup=MAIN_KEYBOARD,
    )

# ---------------------------------------------------------------------------
# Экспорт Excel (5 листов: Наличные поступления/списания, Карта поступления/списания, Запросы)
# ---------------------------------------------------------------------------

async def export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    data    = await asyncio.to_thread(load_data)
    user_id = update.effective_user.id
    txs     = [t for t in data["transactions"] if t["user_id"] == user_id]
    reqs    = [r for r in data["requests"]     if r["user_id"] == user_id]
    trs     = [t for t in data["transfers"]    if t["user_id"] == user_id]

    if not txs and not reqs and not trs:
        await update.message.reply_text("📭 Операций пока нет.", reply_markup=MAIN_KEYBOARD)
        return

    wb = Workbook()
    header_font    = Font(name="Arial", bold=True, color="FFFFFF")
    income_fill    = PatternFill("solid", start_color="1E7E34")
    expense_fill   = PatternFill("solid", start_color="C0392B")
    request_fill   = PatternFill("solid", start_color="2980B9")
    transfer_fill  = PatternFill("solid", start_color="6C3483")

    def style_header(cell, fill):
        cell.font      = header_font
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="center")

    def build_sheet(ws, rows, cols, fill, has_note=False, extra_col_widths: dict | None = None):
        ws.append(cols)
        for i, col in enumerate(cols, 1):
            style_header(ws.cell(1, i), fill)
        for row in rows:
            ws.append(row)
        if rows:
            r = len(rows) + 2
            ws[f"A{r}"] = "Итого"
            ws[f"A{r}"].font = Font(name="Arial", bold=True)
            amount_col = "B"
            ws[f"{amount_col}{r}"] = f"=SUM({amount_col}2:{amount_col}{r-1})"
            ws[f"{amount_col}{r}"].font = Font(name="Arial", bold=True)
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 25
        if has_note:
            ws.column_dimensions["D"].width = 25
        if extra_col_widths:
            for col_letter, width in extra_col_widths.items():
                ws.column_dimensions[col_letter].width = width

    # Лист 1: Наличные — Поступления
    ws1 = wb.active
    ws1.title = "Наличные Поступления"
    rows = [[t["date"], float(t["amount"]), _strip_emoji(t.get("category"))]
            for t in txs if t.get("account") == "cash" and t["type"] == "income"]
    build_sheet(ws1, rows, ["Дата", "Сумма (₽)", "Категория"], income_fill)

    # Лист 2: Наличные — Списания
    ws2 = wb.create_sheet("Наличные Списания")
    rows = [[t["date"], float(t["amount"]), _strip_emoji(t.get("category")), _strip_emoji(t.get("note"))]
            for t in txs if t.get("account") == "cash" and t["type"] == "expense"]
    build_sheet(ws2, rows, ["Дата", "Сумма (₽)", "Категория", "Примечание"], expense_fill, has_note=True)

    # Лист 3: Карта — Поступления
    ws3 = wb.create_sheet("Карта Поступления")
    rows = [[t["date"], float(t["amount"]), _bank_label(t.get("bank"))]
            for t in txs if t.get("account") == "card" and t["type"] == "income"]
    build_sheet(ws3, rows, ["Дата", "Сумма (₽)", "Банк"], income_fill,
                extra_col_widths={"C": 15})

    # Лист 4: Карта — Списания
    ws4 = wb.create_sheet("Карта Списания")
    rows = [[t["date"], float(t["amount"]), _strip_emoji(t.get("category")),
             _strip_emoji(t.get("note")), _bank_label(t.get("bank"))]
            for t in txs if t.get("account") == "card" and t["type"] == "expense"]
    build_sheet(ws4, rows, ["Дата", "Сумма (₽)", "Категория", "Примечание", "Банк"],
                expense_fill, has_note=True, extra_col_widths={"E": 15})

    # Лист 5: Запросы (запрошено, получено по карте, осталось получить)
    ws5 = wb.create_sheet("Запросы")
    req_rows = [[r["date"], float(r["amount"]), _bank_label(r.get("bank"))] for r in reqs]
    build_sheet(ws5, req_rows, ["Дата", "Запрошено (₽)", "Банк"], request_fill,
                extra_col_widths={"C": 15})

    total_card_in = sum(
        Decimal(t["amount"]) for t in txs
        if t.get("account") == "card" and t["type"] == "income"
    )
    if req_rows:
        # Строка "Итого" уже добавлена build_sheet'ом на len(req_rows)+2
        total_r = len(req_rows) + 2
        rec_r   = total_r + 2
        rem_r   = total_r + 3
        ws5[f"A{rec_r}"] = "Получено по карте"
        ws5[f"A{rec_r}"].font = Font(name="Arial", bold=True)
        ws5[f"B{rec_r}"] = float(total_card_in)
        ws5[f"B{rec_r}"].font = Font(name="Arial", bold=True)
        ws5[f"A{rem_r}"] = "Осталось получить"
        ws5[f"A{rem_r}"].font = Font(name="Arial", bold=True)
        ws5[f"B{rem_r}"] = f"=B{total_r}-B{rec_r}"
        ws5[f"B{rem_r}"].font = Font(name="Arial", bold=True)
        ws5.column_dimensions["A"].width = 22

    # Лист 6: Переводы (карта ↔ наличные / карта ↔ карта)
    ws6 = wb.create_sheet("Переводы")

    def _dir_label(t: dict) -> str:
        src = "Карта" if t.get("from") == "card" else "Наличные"
        dst = "Карта" if t.get("to")   == "card" else "Наличные"
        return f"{src} → {dst}"

    tr_rows = [
        [
            t["date"],
            float(t["amount"]),
            _dir_label(t),
            _bank_label(t.get("bank")),
            _bank_label(t.get("bank_to")) if t.get("bank_to") else "",
        ]
        for t in trs
    ]
    build_sheet(ws6, tr_rows, ["Дата", "Сумма (₽)", "Направление", "Банк отправителя", "Банк получателя"],
                transfer_fill, extra_col_widths={"D": 18, "E": 18})

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    now = datetime.now(MOSCOW_TZ).strftime("%Y-%m-%d")
    await update.message.reply_document(
        document=buf,
        filename=f"budget_{now}.xlsx",
        caption="📊 Готово!",
    )

# ---------------------------------------------------------------------------
# Редактирование / удаление
# ---------------------------------------------------------------------------

# --- Вспомогательные функции для отрисовки экранов редактирования ---------

def _edit_item_label(item: dict, is_request: bool, is_transfer: bool = False) -> str:
    if is_request:
        bank = f" [{_bank_label(item.get('bank'))}]" if item.get("bank") else ""
        return f"📨 {fmt(item['amount'])} ₽{bank}   {item['date']}"
    if is_transfer:
        src = "Карта" if item.get("from") == "card" else "Нал"
        dst = "Карта" if item.get("to")   == "card" else "Нал"
        if item.get("from") == "card" and item.get("to") == "card":
            bank_from = f" [{_bank_label(item.get('bank'))}]"    if item.get("bank")    else ""
            bank_dst  = f" [{_bank_label(item.get('bank_to'))}]" if item.get("bank_to") else ""
            return f"🔄 {fmt(item['amount'])} ₽  {src}{bank_from}→{dst}{bank_dst}   {item['date']}"
        bank = f" [{_bank_label(item.get('bank'))}]" if item.get("bank") else ""
        return f"🔄 {fmt(item['amount'])} ₽  {src}{bank}→{dst}   {item['date']}"
    sign = "➕" if item["type"] == "income" else "➖"
    cat  = f" · {item['category']}" if item.get("category") else ""
    note = f" · {item['note']}"     if item.get("note")     else ""
    return f"{sign} {fmt(item['amount'])} ₽{cat}{note}   {item['date']}"


def _edit_get_filtered(data: dict, user_id: int, flt: dict) -> list[dict]:
    if flt["type"] == "requests":
        items = [r for r in data["requests"]  if r["user_id"] == user_id]
    elif flt["type"] == "transfers":
        items = [t for t in data["transfers"] if t["user_id"] == user_id]
    else:
        account   = flt["type"]      # "cash" | "card"
        direction = flt["direction"] # "income" | "expense"
        items = [
            t for t in data["transactions"]
            if t["user_id"] == user_id
            and t.get("account") == account
            and t["type"]       == direction
        ]
    # Свежие сверху
    return items[::-1]


async def _edit_render_type_picker(target):
    """target — это update.message (на старте) или callback_query (после кликов)."""
    buttons = [
        [InlineKeyboardButton("💵 Наличные",  callback_data="edit_type:cash")],
        [InlineKeyboardButton("💳 Карта",     callback_data="edit_type:card")],
        [InlineKeyboardButton("📨 Запросы",   callback_data="edit_type:requests")],
        [InlineKeyboardButton("🔄 Переводы",  callback_data="edit_type:transfers")],
        [InlineKeyboardButton("❌ Отмена",    callback_data="edit_cancel")],
    ]
    text = "Что редактируем?"
    kb   = InlineKeyboardMarkup(buttons)
    if hasattr(target, "edit_message_text"):
        await target.edit_message_text(text, reply_markup=kb)
    else:
        await target.reply_text(text, reply_markup=kb)


async def _edit_render_direction_picker(query, account: str):
    acc_label = "💵 Наличные" if account == "cash" else "💳 Карта"
    buttons = [
        [InlineKeyboardButton("➕ Поступления", callback_data="edit_dir:income")],
        [InlineKeyboardButton("➖ Списания",    callback_data="edit_dir:expense")],
        [
            InlineKeyboardButton("◀️ Назад",  callback_data="edit_back_type"),
            InlineKeyboardButton("❌ Отмена", callback_data="edit_cancel"),
        ],
    ]
    await query.edit_message_text(
        f"{acc_label}\nПоступления или списания?",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def _edit_render_list(query, context, user_id: int):
    flt  = context.user_data["edit_filter"]
    data = await asyncio.to_thread(load_data)
    items = _edit_get_filtered(data, user_id, flt)
    is_request  = flt["type"] == "requests"
    is_transfer = flt["type"] == "transfers"
    # Запросы и переводы возвращают назад к выбору типа; остальные — к выбору направления
    back_cb = "edit_back_type" if (is_request or is_transfer) else "edit_back_dir"

    # Заголовок
    if is_request:
        title = "📨 Запросы"
    elif is_transfer:
        title = "🔄 Переводы"
    else:
        acc  = "💵 Наличные" if flt["type"] == "cash" else "💳 Карта"
        dirn = "Поступления" if flt["direction"] == "income" else "Списания"
        title = f"{acc} → {dirn}"

    if not items:
        buttons = [[
            InlineKeyboardButton("◀️ Назад",  callback_data=back_cb),
            InlineKeyboardButton("❌ Отмена", callback_data="edit_cancel"),
        ]]
        await query.edit_message_text(
            f"{title}\n\nПока нет операций.",
            reply_markup=InlineKeyboardMarkup(buttons),
        )
        return

    total_pages = max(1, (len(items) + EDIT_PAGE_SIZE - 1) // EDIT_PAGE_SIZE)
    page = max(0, min(flt.get("page", 0), total_pages - 1))
    flt["page"] = page

    start = page * EDIT_PAGE_SIZE
    page_items = items[start:start + EDIT_PAGE_SIZE]

    buttons = []
    for it in page_items:
        cb = f"edit_sel:{it['id']}"
        buttons.append([InlineKeyboardButton(
            _edit_item_label(it, is_request, is_transfer), callback_data=cb
        )])

    # Навигация по страницам
    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("◀️", callback_data=f"edit_page:{page - 1}"))
    nav.append(InlineKeyboardButton(f"стр. {page + 1}/{total_pages}", callback_data="edit_noop"))
    if page < total_pages - 1:
        nav.append(InlineKeyboardButton("▶️", callback_data=f"edit_page:{page + 1}"))
    if len(nav) > 1:
        buttons.append(nav)

    buttons.append([
        InlineKeyboardButton("◀️ Назад",  callback_data=back_cb),
        InlineKeyboardButton("❌ Отмена", callback_data="edit_cancel"),
    ])

    await query.edit_message_text(
        f"{title}\nВыбери операцию:",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def _edit_render_item(query, context, item: dict,
                             is_request: bool, is_transfer: bool = False):
    if is_transfer:
        src = "Карта" if item.get("from") == "card" else "Наличные"
        dst = "Карта" if item.get("to")   == "card" else "Наличные"
        if item.get("from") == "card" and item.get("to") == "card":
            bank_line = (
                f"Банк-отправитель: {_bank_label(item.get('bank'))}\n"
                f"Банк-получатель: {_bank_label(item.get('bank_to'))}\n"
            )
        else:
            bank_line = f"Банк: {_bank_label(item.get('bank'))}\n"
        desc = (
            f"🔄 Перевод\n"
            f"Сумма: {fmt(item['amount'])} ₽\n"
            f"{bank_line}"
            f"Направление: {src} → {dst}\n"
            f"Дата: {item['date']}"
        )
        buttons = [
            [InlineKeyboardButton("✏️ Сумму",  callback_data="edit_field:amount")],
            [InlineKeyboardButton("🗑 Удалить", callback_data="edit_delete")],
            [
                InlineKeyboardButton("◀️ Назад",  callback_data="edit_back_list"),
                InlineKeyboardButton("❌ Отмена", callback_data="edit_cancel"),
            ],
        ]
    elif is_request:
        desc = (
            f"📨 Запрос\n"
            f"Сумма: {fmt(item['amount'])} ₽\n"
            f"Банк: {_bank_label(item.get('bank'))}\n"
            f"Дата: {item['date']}"
        )
        buttons = [
            [InlineKeyboardButton("✏️ Сумму",  callback_data="edit_field:amount")],
            [InlineKeyboardButton("🗑 Удалить", callback_data="edit_delete")],
            [
                InlineKeyboardButton("◀️ Назад",  callback_data="edit_back_list"),
                InlineKeyboardButton("❌ Отмена", callback_data="edit_cancel"),
            ],
        ]
    else:
        acc    = "💵 Наличные" if item.get("account") == "cash" else "💳 Карта"
        d_type = "Поступление" if item["type"] == "income" else "Списание"
        cat    = item.get("category") or "—"
        note   = item.get("note") or "—"
        bank_line = f"Банк: {_bank_label(item.get('bank'))}\n" if item.get("account") == "card" else ""
        desc = (
            f"{acc} | {d_type}\n"
            f"Сумма: {fmt(item['amount'])} ₽\n"
            f"{bank_line}"
            f"Категория: {cat}\n"
            f"Примечание: {note}\n"
            f"Дата: {item['date']}"
        )
        buttons = [
            [InlineKeyboardButton("✏️ Сумму",      callback_data="edit_field:amount")],
            [InlineKeyboardButton("✏️ Категорию",  callback_data="edit_field:category")],
            [InlineKeyboardButton("✏️ Примечание", callback_data="edit_field:note")],
            [InlineKeyboardButton("🗑 Удалить",    callback_data="edit_delete")],
            [
                InlineKeyboardButton("◀️ Назад",  callback_data="edit_back_list"),
                InlineKeyboardButton("❌ Отмена", callback_data="edit_cancel"),
            ],
        ]
    await query.edit_message_text(desc, reply_markup=InlineKeyboardMarkup(buttons))


async def _edit_render_delete_confirm(query, item: dict,
                                       is_request: bool, is_transfer: bool = False):
    if is_transfer:
        src = "Карта" if item.get("from") == "card" else "Наличные"
        dst = "Карта" if item.get("to")   == "card" else "Наличные"
        if item.get("from") == "card" and item.get("to") == "card":
            bank_from = f" [{_bank_label(item.get('bank'))}]"    if item.get("bank")    else ""
            bank_dst  = f" [{_bank_label(item.get('bank_to'))}]" if item.get("bank_to") else ""
            text = (f"Удалить перевод {fmt(item['amount'])} ₽  "
                    f"{src}{bank_from} → {dst}{bank_dst}  от {item['date']}?")
        else:
            bank = f" [{_bank_label(item.get('bank'))}]" if item.get("bank") else ""
            text = (f"Удалить перевод {fmt(item['amount'])} ₽  "
                    f"{src}{bank} → {dst}  от {item['date']}?")
    elif is_request:
        text = f"Удалить запрос {fmt(item['amount'])} ₽ от {item['date']}?"
    else:
        sign = "+" if item["type"] == "income" else "−"
        cat  = f" · {item['category']}" if item.get("category") else ""
        text = f"Удалить операцию {sign}{fmt(item['amount'])} ₽{cat} от {item['date']}?"
    buttons = [[
        InlineKeyboardButton("🗑 Да, удалить", callback_data="edit_confirm_yes"),
        InlineKeyboardButton("↩️ Нет",         callback_data="edit_confirm_no"),
    ]]
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(buttons))


# --- Обработчики ----------------------------------------------------------

async def edit_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Точка входа: команда /edit или кнопка «✏️ Изменить»."""
    context.user_data.pop("edit_filter",      None)
    context.user_data.pop("edit_tx_id",       None)
    context.user_data.pop("edit_is_request",  None)
    context.user_data.pop("edit_is_transfer", None)
    context.user_data.pop("edit_field",       None)
    await _edit_render_type_picker(update.message)
    return EDIT_CHOOSE_TYPE


async def edit_pick_type(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "edit_cancel":
        await query.edit_message_text("❌ Отменено.")
        return ConversationHandler.END

    if not query.data.startswith("edit_type:"):
        return EDIT_CHOOSE_TYPE

    typ = query.data.split(":", 1)[1]   # "cash" | "card" | "requests" | "transfers"
    context.user_data["edit_filter"] = {"type": typ, "page": 0}

    if typ in ("requests", "transfers"):
        await _edit_render_list(query, context, update.effective_user.id)
        return EDIT_LIST

    await _edit_render_direction_picker(query, typ)
    return EDIT_CHOOSE_DIRECTION


async def edit_pick_direction(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "edit_cancel":
        await query.edit_message_text("❌ Отменено.")
        return ConversationHandler.END

    if query.data == "edit_back_type":
        await _edit_render_type_picker(query)
        return EDIT_CHOOSE_TYPE

    if not query.data.startswith("edit_dir:"):
        return EDIT_CHOOSE_DIRECTION

    direction = query.data.split(":", 1)[1]
    context.user_data["edit_filter"]["direction"] = direction
    context.user_data["edit_filter"]["page"] = 0

    await _edit_render_list(query, context, update.effective_user.id)
    return EDIT_LIST


async def edit_list_action(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "edit_cancel":
        await query.edit_message_text("❌ Отменено.")
        return ConversationHandler.END

    if query.data == "edit_noop":
        return EDIT_LIST

    if query.data == "edit_back_type":
        await _edit_render_type_picker(query)
        return EDIT_CHOOSE_TYPE

    if query.data == "edit_back_dir":
        flt = context.user_data["edit_filter"]
        await _edit_render_direction_picker(query, flt["type"])
        return EDIT_CHOOSE_DIRECTION

    if query.data.startswith("edit_page:"):
        page = int(query.data.split(":", 1)[1])
        context.user_data["edit_filter"]["page"] = page
        await _edit_render_list(query, context, update.effective_user.id)
        return EDIT_LIST

    if query.data.startswith("edit_sel:"):
        item_id = query.data.split(":", 1)[1]
        flt     = context.user_data["edit_filter"]
        is_req  = flt["type"] == "requests"
        is_tr   = flt["type"] == "transfers"
        data    = await asyncio.to_thread(load_data)
        if is_req:
            coll = data["requests"]
        elif is_tr:
            coll = data["transfers"]
        else:
            coll = data["transactions"]
        item = next((x for x in coll if x["id"] == item_id), None)
        if item is None or item["user_id"] != update.effective_user.id:
            await query.edit_message_text("⛔ Операция недоступна.")
            return ConversationHandler.END
        context.user_data["edit_tx_id"]       = item_id
        context.user_data["edit_is_request"]  = is_req
        context.user_data["edit_is_transfer"] = is_tr
        await _edit_render_item(query, context, item, is_req, is_tr)
        return EDIT_CHOOSE_FIELD

    return EDIT_LIST


async def edit_field_action(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "edit_cancel":
        await query.edit_message_text("❌ Отменено.")
        return ConversationHandler.END

    if query.data == "edit_back_list":
        await _edit_render_list(query, context, update.effective_user.id)
        return EDIT_LIST

    if query.data == "edit_delete":
        tx_id  = context.user_data.get("edit_tx_id")
        is_req = context.user_data.get("edit_is_request", False)
        is_tr  = context.user_data.get("edit_is_transfer", False)
        data   = await asyncio.to_thread(load_data)
        if is_req:
            coll = data["requests"]
        elif is_tr:
            coll = data["transfers"]
        else:
            coll = data["transactions"]
        item = next((x for x in coll if x["id"] == tx_id), None)
        if item is None or item["user_id"] != update.effective_user.id:
            await query.edit_message_text("⛔ Операция недоступна.")
            return ConversationHandler.END
        await _edit_render_delete_confirm(query, item, is_req, is_tr)
        return EDIT_CONFIRM_DELETE

    if query.data.startswith("edit_field:"):
        field  = query.data.split(":", 1)[1]
        is_req = context.user_data.get("edit_is_request", False)
        is_tr  = context.user_data.get("edit_is_transfer", False)
        # У запросов и переводов редактируется только сумма
        if (is_req or is_tr) and field != "amount":
            return EDIT_CHOOSE_FIELD
        context.user_data["edit_field"] = field
        prompts = {
            "amount":   "Введи новую сумму:",
            "category": "Введи новую категорию:",
            "note":     "Введи новое примечание (или /skip для удаления):",
        }
        await query.edit_message_text(prompts.get(field, "Введи значение:"))
        return EDIT_ENTERING_VALUE

    return EDIT_CHOOSE_FIELD


async def edit_confirm_delete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "edit_confirm_no":
        # Возврат к карточке операции
        tx_id  = context.user_data.get("edit_tx_id")
        is_req = context.user_data.get("edit_is_request", False)
        is_tr  = context.user_data.get("edit_is_transfer", False)
        data   = await asyncio.to_thread(load_data)
        if is_req:
            coll = data["requests"]
        elif is_tr:
            coll = data["transfers"]
        else:
            coll = data["transactions"]
        item = next((x for x in coll if x["id"] == tx_id), None)
        if item is None or item["user_id"] != update.effective_user.id:
            await query.edit_message_text("⛔ Операция недоступна.")
            return ConversationHandler.END
        await _edit_render_item(query, context, item, is_req, is_tr)
        return EDIT_CHOOSE_FIELD

    if query.data == "edit_confirm_yes":
        tx_id   = context.user_data.get("edit_tx_id")
        is_req  = context.user_data.get("edit_is_request", False)
        is_tr   = context.user_data.get("edit_is_transfer", False)
        user_id = update.effective_user.id
        if is_req:
            key = "requests"
        elif is_tr:
            key = "transfers"
        else:
            key = "transactions"

        deleted_amount = [None]

        def mutate(data):
            item = next((x for x in data[key] if x.get("id") == tx_id), None)
            if item is None or item.get("user_id") != user_id:
                return
            deleted_amount[0] = item["amount"]
            data[key] = [x for x in data[key] if x.get("id") != tx_id]

        try:
            await _atomic_modify(mutate)
        except OSError:
            logger.exception("Failed to delete item")
            await query.edit_message_text("⚠️ Не удалось удалить, попробуй ещё раз.")
            return ConversationHandler.END

        if deleted_amount[0] is None:
            await query.edit_message_text("⛔ Операция недоступна.")
            return ConversationHandler.END

        if is_tr:
            msg = f"🗑 Перевод {fmt(deleted_amount[0])} ₽ удалён."
        elif is_req:
            msg = f"🗑 Запрос {fmt(deleted_amount[0])} ₽ удалён."
        else:
            msg = f"🗑 Операция {fmt(deleted_amount[0])} ₽ удалена."
        await query.edit_message_text(msg)
        return ConversationHandler.END

    return EDIT_CONFIRM_DELETE


async def edit_receive_value(update: Update, context: ContextTypes.DEFAULT_TYPE):
    field   = context.user_data.get("edit_field")
    tx_id   = context.user_data.get("edit_tx_id")
    is_req  = context.user_data.get("edit_is_request", False)
    text    = (update.message.text or "").strip()
    user_id = update.effective_user.id

    # Сначала валидируем ввод — без обращения к данным
    if field == "amount":
        try:
            new_val = parse_amount(text)
        except ValueError:
            await update.message.reply_text("❌ Введи корректную сумму:")
            return EDIT_ENTERING_VALUE
        new_value = str(new_val)
    elif field == "category":
        if not text or text == "/skip":
            await update.message.reply_text("❌ Категория не может быть пустой:")
            return EDIT_ENTERING_VALUE
        new_value = text[:MAX_NOTE_LEN]
    elif field == "note":
        if text == "/skip":
            new_value = None
        elif not text:
            await update.message.reply_text("❌ Введи примечание (или /skip для удаления):")
            return EDIT_ENTERING_VALUE
        else:
            new_value = text[:MAX_NOTE_LEN]
    else:
        await update.message.reply_text("⛔ Неизвестное поле.", reply_markup=MAIN_KEYBOARD)
        return ConversationHandler.END

    is_tr = context.user_data.get("edit_is_transfer", False)
    if is_req:
        key = "requests"
    elif is_tr:
        key = "transfers"
    else:
        key = "transactions"
    found = [False]

    def mutate(data):
        item = next((x for x in data[key] if x.get("id") == tx_id), None)
        if item is None or item.get("user_id") != user_id:
            return
        item[field] = new_value
        found[0] = True

    try:
        await _atomic_modify(mutate)
    except OSError:
        logger.exception("Failed to save edited value")
        await update.message.reply_text("⚠️ Не удалось сохранить, попробуй ещё раз.", reply_markup=MAIN_KEYBOARD)
        return ConversationHandler.END

    if not found[0]:
        await update.message.reply_text("⛔ Операция недоступна.", reply_markup=MAIN_KEYBOARD)
        return ConversationHandler.END

    if field == "amount":
        msg = f"✅ Сумма обновлена: {fmt(new_value)} ₽"
    elif field == "category":
        msg = f"✅ Категория обновлена: {new_value}"
    elif new_value is None:
        msg = "✅ Примечание удалено."
    else:
        msg = f"✅ Примечание обновлено: {new_value}"

    await update.message.reply_text(msg, reply_markup=MAIN_KEYBOARD)
    return ConversationHandler.END

# ---------------------------------------------------------------------------
# /clear
# ---------------------------------------------------------------------------

async def clear(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton("🗑 Да, удалить всё", callback_data="clear_yes"),
        InlineKeyboardButton("Отмена",             callback_data="clear_no"),
    ]])
    await update.message.reply_text("Точно удалить ВСЕ свои операции? Это необратимо.", reply_markup=kb)


async def clear_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "clear_yes":
        user_id = update.effective_user.id

        def mutate(data):
            data["transactions"] = [t for t in data["transactions"] if t.get("user_id") != user_id]
            data["requests"]     = [r for r in data["requests"]     if r.get("user_id") != user_id]
            data["transfers"]    = [t for t in data["transfers"]    if t.get("user_id") != user_id]

        try:
            await _atomic_modify(mutate)
        except OSError:
            logger.exception("Failed to clear user data")
            await query.edit_message_text("⚠️ Не удалось очистить, попробуй ещё раз.")
            return
        await query.edit_message_text("🗑 Все твои данные удалены.")
    else:
        await query.edit_message_text("Отменено.")


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text("❌ Отменено.", reply_markup=MAIN_KEYBOARD)
    return ConversationHandler.END

# ---------------------------------------------------------------------------
# Запуск
# ---------------------------------------------------------------------------

def main():
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    if not token:
        raise ValueError("Установи переменную окружения TELEGRAM_BOT_TOKEN")

    proxy_url = os.environ.get("PROXY_URL")
    builder = (
        Application.builder()
        .token(token)
        .get_updates_read_timeout(25)
        .get_updates_write_timeout(25)
        .get_updates_connect_timeout(10)
        .get_updates_pool_timeout(25)
    )
    if proxy_url:
        builder = builder.proxy(proxy_url).get_updates_proxy(proxy_url)
    app = builder.build()

    app.add_error_handler(on_error)

    main_filter = filters.Regex(
        "^(💵 Наличные|💳 Карта|💰 Баланс|🕓 История|📨 Запросил|🔄 Перевод)$"
    )

    add_conv = ConversationHandler(
        entry_points=[
            CommandHandler("start", start),
            MessageHandler(main_filter, handle_account),
        ],
        states={
            ST_CHOOSE_BANK:              [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_bank)],
            ST_CHOOSE_DIRECTION:         [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_direction)],
            ST_ENTERING_AMOUNT:          [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_amount)],
            ST_CHOOSE_CATEGORY:          [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_category)],
            ST_ENTERING_ZP_DATE:         [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_zp_date)],
            ST_ENTERING_NOTE:            [CommandHandler("skip", handle_note), MessageHandler(filters.TEXT & ~filters.COMMAND, handle_note)],
            ST_ENTERING_REQUEST_AMOUNT:  [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_request_amount)],
            ST_CHOOSE_TRANSFER_DIR:      [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_transfer_dir)],
            ST_CHOOSE_TRANSFER_BANK_TO:  [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_transfer_bank_to)],
            ST_ENTERING_TRANSFER_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_transfer_amount)],
        },
        fallbacks=[
            CommandHandler("cancel", cancel),
            CommandHandler("start", start),
        ],
    )

    edit_conv = ConversationHandler(
        entry_points=[
            CommandHandler("edit", edit_start),
            MessageHandler(filters.Regex("^✏️ Изменить$"), edit_start),
        ],
        allow_reentry=True,
        states={
            EDIT_CHOOSE_TYPE: [
                CallbackQueryHandler(edit_pick_type, pattern="^(edit_type:|edit_cancel$)"),
            ],
            EDIT_CHOOSE_DIRECTION: [
                CallbackQueryHandler(edit_pick_direction, pattern="^(edit_dir:|edit_back_type$|edit_cancel$)"),
            ],
            EDIT_LIST: [
                CallbackQueryHandler(
                    edit_list_action,
                    pattern="^(edit_sel:|edit_page:|edit_back_type$|edit_back_dir$|edit_noop$|edit_cancel$)",
                ),
            ],
            EDIT_CHOOSE_FIELD: [
                CallbackQueryHandler(
                    edit_field_action,
                    pattern="^(edit_field:|edit_delete$|edit_back_list$|edit_cancel$)",
                ),
            ],
            EDIT_CONFIRM_DELETE: [
                CallbackQueryHandler(edit_confirm_delete, pattern="^edit_confirm_(yes|no)$"),
            ],
            EDIT_ENTERING_VALUE: [
                CommandHandler("skip", edit_receive_value),
                MessageHandler(filters.TEXT & ~filters.COMMAND, edit_receive_value),
            ],
        },
        fallbacks=[
            CommandHandler("cancel", cancel),
            CommandHandler("start", start),
        ],
    )

    app.add_handler(CommandHandler("export", export_excel))
    app.add_handler(CommandHandler("clear",  clear))
    app.add_handler(CallbackQueryHandler(clear_confirm, pattern="^clear_"))
    # edit_conv ставим первым: чтобы кнопка «✏️ Изменить» захватывалась им,
    # а не общей add_conv (которая иначе съест регекспом из main_filter).
    app.add_handler(edit_conv)
    app.add_handler(add_conv)

    logger.info("Bot started")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
